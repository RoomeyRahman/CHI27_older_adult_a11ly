#!/usr/bin/env python3
"""Build the CHI knowledge table workbook + markdown from extraction JSONs.

Usage:
    python build_matrix.py work/ --bib work/bib_index.json \
        --out out.xlsx --md out.md
    python build_matrix.py --schema      # print the synthesis.json schema

Expected inside <workdir>:
    extractions/*.json        one record per paper (references/extraction-fields.md)
    synthesis.json            final knowledge, conflicts, gaps (--schema)
    consolidation_map.json    candidate themes -> knowledge id (optional, merged
                              from synthesis.json if present there instead)

The script interprets nothing. All judgment lives in the JSON. Its one active
job is validating that every citation key resolves against the bib index, so
unresolved keys surface loudly instead of reaching the manuscript.
"""

import argparse
import json
import re
import sys
from pathlib import Path

MISSING = "Not reported"
CITE_RE = re.compile(r"\\cite\w*\{([^}]*)\}")

SYNTHESIS_SCHEMA = r"""
synthesis.json
{
  "topic": "one-sentence topic statement",
  "focus": "the focus instruction this run was conducted under",
  "year_range": "2018-2026",
  "max_knowledge": 10,
  "knowledge": [
    {
      "id": "K1",
      "statement": "Declarative claim the field would endorse. No study named.",
      "lineage": "Established in lab settings \\cite{smith2021}; replicated with clinicians \\cite{lee2022}; extended to voice-only delivery \\cite{okafor2023}; bounded to high-stakes tasks \\cite{chen2024}.",
      "evidence": "4 studies; 3 lab, 1 field; Moderate - convergent but single-session",
      "gap": "Gap: no study tested longitudinal recalibration -> author to state contribution",
      "bib_keys": ["smith2021", "lee2022", "okafor2023", "chen2024"]
    }
  ],
  "consolidation": [
    {
      "knowledge_id": "K1",
      "candidate_themes": ["explanation confidence and reliance", "verbosity effects on trust"],
      "paper_ids": ["smith2021", "lee2022"],
      "reason": "Same construct; Lee extends Smith to clinicians."
    }
  ],
  "conflicts": [
    {
      "claim": "Adaptive notifications reduce perceived interruption",
      "supporting": ["smith2021"], "contradicting": ["okafor2022"],
      "explanation": "Supporting studies used self-report; the contradicting one logged dismissals.",
      "confidence": "Moderate"
    }
  ],
  "gaps": [
    {"rank": 1, "gap": "...", "knowledge_id": "K1", "known": "...", "unknown": "...",
     "why_it_matters": "...", "research_question": "...", "method": "...",
     "originality": "4/5 - ...", "feasibility": "5/5 - ..."}
  ],
  "discussion_order": "<=150 words on the order to present K1..Kn and why"
}
"""


# ----------------------------------------------------------------- helpers

def g(rec, *keys):
    cur = rec
    for k in keys:
        if not isinstance(cur, dict) or k not in cur or cur[k] in (None, "", []):
            return MISSING
        cur = cur[k]
    return str(cur)


def join(*parts, sep="; "):
    vals = [p for p in parts if p and p != MISSING]
    return sep.join(vals) if vals else MISSING


def listjoin(val, sep="; "):
    if not val:
        return MISSING
    if isinstance(val, str):
        return val
    return sep.join(str(v) for v in val)


def cite_short(rec):
    authors = g(rec, "citation", "authors")
    year = g(rec, "citation", "year")
    if authors == MISSING:
        return f"{rec.get('paper_id', 'unknown')} ({year})"
    first = authors.split(",")[0].strip()
    multi = "," in authors or "&" in authors or " and " in authors
    return f"{first}{' et al.' if multi else ''} ({year})"


def sample_str(rec):
    s = rec.get("sample") or {}
    return f"N={s.get('n', MISSING)}; {s.get('population', MISSING)}"


def findings_str(rec):
    out = []
    for f in rec.get("key_findings") or []:
        bits = [str(f.get("claim", "")).strip()]
        if f.get("evidence"):
            bits.append(f"[{f['evidence']}]")
        if f.get("page"):
            bits.append(f"({f['page']})")
        out.append(" ".join(b for b in bits if b))
    return "\n".join(out) if out else MISSING


MATRIX_COLS = [
    ("Cite key", lambda r: r.get("bib_key") or r.get("paper_id", MISSING)),
    ("Study", cite_short),
    ("Year", lambda r: g(r, "citation", "year")),
    ("Venue", lambda r: g(r, "citation", "venue")),
    ("Country/Context", lambda r: g(r, "sample", "country")),
    ("Research question", lambda r: g(r, "research_question")),
    ("Theory", lambda r: g(r, "theory")),
    ("Sample", sample_str),
    ("Method", lambda r: join(g(r, "study_type"), g(r, "analysis"))),
    ("Measures", lambda r: listjoin(r.get("measures"))),
    ("Main finding", findings_str),
    ("Stated limitation", lambda r: listjoin(r.get("author_limitations"))),
    ("Knowledge ID", lambda r: listjoin(r.get("knowledge_ids") or r.get("theme_tags"))),
    ("Relevance", lambda r: g(r, "relevance_to_topic")),
]


def load(workdir):
    wd = Path(workdir)
    ext_dir = wd / "extractions" if (wd / "extractions").is_dir() else wd
    records, syn, consol = [], None, None
    for p in sorted(ext_dir.glob("*.json")):
        if p.name in ("synthesis.json", "consolidation_map.json", "bib_index.json",
                      "candidate_themes.json"):
            continue
        records.append(json.loads(p.read_text(encoding="utf-8")))
    for name, target in (("synthesis.json", "syn"), ("consolidation_map.json", "consol")):
        for cand in (wd / name, ext_dir / name):
            if cand.is_file():
                data = json.loads(cand.read_text(encoding="utf-8"))
                if target == "syn":
                    syn = data
                else:
                    consol = data
                break
    return records, syn, consol


def collect_keys(syn, records):
    """Every citation key referenced anywhere in the synthesis."""
    keys = set()
    for k in (syn or {}).get("knowledge", []) or (syn or {}).get("themes", []):
        keys.update(k.get("bib_keys") or [])
        keys.update(CITE_RE.findall(k.get("lineage", "")) and
                    [x.strip() for grp in CITE_RE.findall(k.get("lineage", ""))
                     for x in grp.split(",") if x.strip()])
    for c in (syn or {}).get("conflicts", []):
        keys.update(c.get("supporting") or [])
        keys.update(c.get("contradicting") or [])
    for r in records:
        if r.get("bib_key"):
            keys.add(r["bib_key"])
    return keys


def validate_keys(keys, bib_path):
    if not bib_path:
        return [], None
    idx = json.loads(Path(bib_path).read_text(encoding="utf-8"))
    known = set(idx.get("entries", {}))
    return sorted(k for k in keys if k not in known), idx


# ----------------------------------------------------------------- tables

def table_knowledge(syn):
    rows = [["ID", "Established knowledge (what the field knows)",
             "How it was built (lineage across studies)",
             "Evidence & consensus", "Gap → extension for this work"]]
    items = syn.get("knowledge") or syn.get("themes") or []
    for i, k in enumerate(items, 1):
        rows.append([k.get("id", f"K{i}"),
                     k.get("statement") or k.get("established", MISSING),
                     k.get("lineage", MISSING),
                     k.get("evidence", MISSING),
                     k.get("gap", MISSING)])
    return rows


def table_consolidation(syn, consol):
    entries = (consol or {}).get("consolidation") or (syn or {}).get("consolidation") or []
    rows = [["Candidate theme", "Papers", "Merged into", "Reason for merge"]]
    for c in entries:
        themes = c.get("candidate_themes") or [c.get("candidate_theme", MISSING)]
        for t in themes:
            rows.append([t, listjoin(c.get("paper_ids")),
                         c.get("knowledge_id", MISSING), c.get("reason", MISSING)])
    return rows


def table_matrix(records):
    rows = [[c[0] for c in MATRIX_COLS]]
    for r in sorted(records, key=lambda x: str(g(x, "citation", "year"))):
        row = [fn(r) for _, fn in MATRIX_COLS]
        if r.get("flags"):
            row[1] += " ⚠"
        rows.append(row)
    return rows


def table_conflicts(syn):
    rows = [["Claim", "Studies supporting", "Studies contradicting",
             "Plausible reason for divergence", "Confidence"]]
    for c in syn.get("conflicts", []):
        rows.append([c.get("claim", MISSING), listjoin(c.get("supporting")),
                     listjoin(c.get("contradicting")),
                     c.get("explanation", MISSING), c.get("confidence", MISSING)])
    return rows


def table_gaps(syn):
    rows = [["Rank", "Gap", "Knowledge ID", "What is known", "What is unknown",
             "Why it matters", "Candidate research question", "Feasible method",
             "Originality", "Feasibility"]]
    for gp in sorted(syn.get("gaps", []), key=lambda x: x.get("rank", 99)):
        rows.append([str(gp.get("rank", MISSING)), gp.get("gap", MISSING),
                     gp.get("knowledge_id", MISSING), gp.get("known", MISSING),
                     gp.get("unknown", MISSING), gp.get("why_it_matters", MISSING),
                     gp.get("research_question", MISSING), gp.get("method", MISSING),
                     gp.get("originality", MISSING), gp.get("feasibility", MISSING)])
    return rows


# ----------------------------------------------------------------- writers

WIDTHS = {
    "0 Final knowledge": [8, 58, 58, 30, 44],
    "1 Consolidation map": [38, 30, 14, 60],
}


def write_xlsx(tables, path, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1F3B57")
    font = Font(bold=True, color="FFFFFF", size=11)

    for name, rows in tables:
        ws = wb.create_sheet(name[:31])
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.fill, cell.font = fill, font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32
        n = len(rows[0])
        widths = WIDTHS.get(name, [26] * n)
        for i in range(n):
            ws.column_dimensions[get_column_letter(i + 1)].width = widths[i % len(widths)]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"

    ws = wb.create_sheet("Run info")
    for k, v in meta.items():
        ws.append([k, str(v)])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows():
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def md_escape(s):
    return str(s).replace("|", "\\|").replace("\n", "<br>")


def write_md(tables, path, syn, meta):
    out = ["# Established knowledge synthesis\n"]
    for k in ("topic", "focus", "year_range"):
        if syn and syn.get(k):
            out.append(f"**{k.replace('_', ' ').title()}:** {syn[k]}  ")
    out.append(f"\n_{meta['papers']} papers · {meta['knowledge_rows']} knowledge statements_\n")
    for name, rows in tables:
        out.append(f"\n## {name}\n")
        out.append("| " + " | ".join(md_escape(c) for c in rows[0]) + " |")
        out.append("|" + "---|" * len(rows[0]))
        for r in rows[1:]:
            out.append("| " + " | ".join(md_escape(c) for c in r) + " |")
    if syn and syn.get("discussion_order"):
        out.append("\n## Suggested order for the write-up\n")
        out.append(syn["discussion_order"])
    if meta.get("unresolved_keys"):
        out.append("\n## ⚠ Unresolved citation keys\n")
        out.append("These keys are not in the .bib index and must be fixed before use:\n")
        out.append(", ".join(f"`{k}`" for k in meta["unresolved_keys"]))
    Path(path).write_text("\n".join(out), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workdir", nargs="?", help="dir with extractions/ + synthesis.json")
    ap.add_argument("--bib", help="bib_index.json from bib_index.py")
    ap.add_argument("--out", help="output .xlsx path")
    ap.add_argument("--md", help="output .md path")
    ap.add_argument("--schema", action="store_true")
    ap.add_argument("--strict", action="store_true",
                    help="exit nonzero if any citation key is unresolved")
    args = ap.parse_args()

    if args.schema:
        print(SYNTHESIS_SCHEMA)
        return
    if not args.workdir:
        ap.error("workdir is required unless --schema is given")

    records, syn, consol = load(args.workdir)
    if not records:
        sys.exit(f"No extraction JSONs found under {args.workdir}")

    unresolved, _ = validate_keys(collect_keys(syn, records), args.bib)

    tables = []
    if syn:
        tables.append(("0 Final knowledge", table_knowledge(syn)))
        cons = table_consolidation(syn, consol)
        if len(cons) > 1:
            tables.append(("1 Consolidation map", cons))
    tables.append(("2 Evidence matrix", table_matrix(records)))
    if syn:
        tables.append(("3 Agreement & conflict", table_conflicts(syn)))
        tables.append(("4 Ranked gaps", table_gaps(syn)))

    k_rows = len(table_knowledge(syn)) - 1 if syn else 0
    meta = {
        "topic": (syn or {}).get("topic", MISSING),
        "focus": (syn or {}).get("focus", MISSING),
        "year_range": (syn or {}).get("year_range", MISSING),
        "papers": len(records),
        "knowledge_rows": k_rows,
        "reduction": f"{len(records)} papers → {k_rows} knowledge statements" if k_rows else "n/a",
        "flagged_studies": ", ".join(r.get("paper_id", "?") for r in records if r.get("flags")) or "none",
        "unresolved_keys": unresolved,
    }

    if args.out:
        write_xlsx(tables, args.out, meta)
        print(f"wrote {args.out}")
    if args.md:
        write_md(tables, args.md, syn, meta)
        print(f"wrote {args.md}")

    print(meta["reduction"])
    if k_rows and k_rows > (syn or {}).get("max_knowledge", 12):
        print(f"WARNING: {k_rows} knowledge rows exceeds max_knowledge — consolidate further.")
    if unresolved:
        print(f"UNRESOLVED CITATION KEYS ({len(unresolved)}): {', '.join(unresolved)}")
        if args.strict:
            sys.exit(1)


if __name__ == "__main__":
    main()
